from tf_idf_model import Similarity
import pandas as pd
from openpyxl import Workbook
wb=Workbook()    #创建文件对象
ws=wb.create_sheet("Mysheet")     #获取第一个sheet
ws.title="teacher"
ws['A1']="query文本"
ws['B1']="教师id "
ws['C1']="专利id "
ws['D1']="教师姓名 "
ws['E1']="专利摘要 "
wb.save(r'F:\\projects\\match_gz\\sample.xlsx')

def save_txt():
    s=Similarity()
    samp1=s.send_query('人工智能')
    samp2=s.send_query('大数据')
    samp3=s.send_query('并行计算')
    samp4=s.send_query('数据挖掘')
    samp5=s.send_query('AI')
    samp6=s.send_query('模式识别')
    samp7=s.send_query('语音处理')
    samp8=s.send_query('自动化')
    samp9=s.send_query('软件测试')
    samp10=s.send_query('云计算')
    file = open(r'query.txt', 'w')
    file.write(samp1 + '\n')
    file.write(samp2 + '\n')
    file.write(samp3 + '\n')
    file.write(samp4 + '\n')
    file.write(samp5 + '\n')
    file.write(samp6 + '\n')
    file.write(samp7 + '\n')
    file.write(samp8 + '\n')
    file.write(samp9 + '\n')
    file.write(samp10 + '\n')

